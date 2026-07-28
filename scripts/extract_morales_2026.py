#!/usr/bin/env python3
"""Extract Morales et al. (2026) Kd worksheet to verified long format."""

from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source" / "Morales_2026_SI.xlsx"
OUTPUT = ROOT / "data" / "source" / "morales_long_reextracted.csv"


def extract_morales(source: Path) -> pd.DataFrame:
    formula_book = load_workbook(source, data_only=False, read_only=False)
    value_book = load_workbook(source, data_only=True, read_only=False)
    formula_sheet = formula_book["Kd"]
    value_sheet = value_book["Kd"]

    blocks: list[tuple[int, str]] = []
    for column in range(1, formula_sheet.max_column + 1):
        label = formula_sheet.cell(7, column).value
        if isinstance(label, str) and label.strip().lower().startswith("kd "):
            blocks.append((column, label.strip()[3:]))

    if len(blocks) != 14:
        raise ValueError(f"Expected 14 PFAS blocks, found {len(blocks)}: {blocks}")

    records: list[dict[str, object]] = []
    for row in range(11, value_sheet.max_row + 1):
        lab_id = value_sheet.cell(row, 1).value
        paper_id = value_sheet.cell(row, 2).value
        if lab_id is None or paper_id is None:
            continue

        for kd_column, pfas in blocks:
            soil = value_sheet.cell(row, kd_column - 2).value
            leachate = value_sheet.cell(row, kd_column - 1).value
            kd_l_per_g = value_sheet.cell(row, kd_column).value
            log_kd = value_sheet.cell(row, kd_column + 1).value

            if not (
                isinstance(kd_l_per_g, (int, float))
                and kd_l_per_g > 0
                and isinstance(log_kd, (int, float))
            ):
                continue

            recalculated_log = math.log10(kd_l_per_g * 1000.0)
            if not math.isclose(log_kd, recalculated_log, abs_tol=1e-10):
                raise ValueError(
                    f"Cached log Kd mismatch at row {row}, {pfas}: "
                    f"cached={log_kd}, recalculated={recalculated_log}"
                )

            records.append(
                {
                    "lab_id": str(lab_id).strip(),
                    "paper_id": str(paper_id).strip(),
                    "OC_pct": value_sheet.cell(row, 4).value,
                    "pH": value_sheet.cell(row, 5).value,
                    "PFAS": pfas,
                    "soil_conc_ug_kg": soil,
                    "leachate_conc_ng_L": leachate,
                    "Kd_L_per_g": kd_l_per_g,
                    "log_Kd_L_per_kg": log_kd,
                    "excel_row": row,
                }
            )

    result = pd.DataFrame.from_records(records)
    if len(result) != 362:
        raise ValueError(f"Expected 362 measured Kd records, found {len(result)}")
    if result.duplicated(["lab_id", "paper_id", "PFAS"]).any():
        raise ValueError("Duplicate soil-PFAS records found")
    return result


if __name__ == "__main__":
    extracted = extract_morales(SOURCE)
    extracted.to_csv(OUTPUT, index=False)
    print(f"Wrote {len(extracted)} rows × {extracted['PFAS'].nunique()} PFAS to {OUTPUT}")
    print(
        "log10 Kd range: "
        f"{extracted['log_Kd_L_per_kg'].min():.6f} to "
        f"{extracted['log_Kd_L_per_kg'].max():.6f}"
    )
